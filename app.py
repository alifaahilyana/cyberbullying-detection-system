"""
Web-Based Cyberbullying Detection and Emotion Analysis System
Final Year Project - Streamlit Application
"""

import re
import io
import joblib
import numpy as np
import pandas as pd
import streamlit as st
import nltk
nltk.download('punkt')
nltk.download('stopwords')
nltk.download('punkt_tab') 
import plotly.express as px
import plotly.graph_objects as go
from nltk.tokenize import word_tokenize
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────────────────────
# PAGE CONFIG
# ──────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="CyberGuard — Cyberbullying Analysis",
    page_icon="🛡️",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ──────────────────────────────────────────────────────────────────────────────
# GLOBAL CSS  (dark-navy + electric-teal palette)
# ──────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=Space+Grotesk:wght@500;700&display=swap');

/* ── Reset & base ── */
html, body, [class*="css"] { font-family: 'Inter', sans-serif; }

.stApp {
    background: #0d1117;
    color: #e6edf3;
}

/* ── Header ── */
.cg-header {
    background: linear-gradient(135deg, #0d1117 0%, #161b22 60%, #0d2137 100%);
    border-bottom: 1px solid #21262d;
    padding: 2rem 2.5rem 1.5rem;
    margin-bottom: 2rem;
}
.cg-header h1 {
    font-family: 'Space Grotesk', sans-serif;
    font-size: 2rem;
    font-weight: 700;
    color: #e6edf3;
    margin: 0 0 0.25rem;
    letter-spacing: -0.5px;
}
.cg-header p {
    color: #8b949e;
    font-size: 0.9rem;
    margin: 0;
}
.cg-badge {
    display: inline-block;
    background: rgba(0, 200, 200, 0.12);
    color: #00c8c8;
    border: 1px solid rgba(0, 200, 200, 0.3);
    padding: 0.2rem 0.7rem;
    border-radius: 999px;
    font-size: 0.75rem;
    font-weight: 600;
    letter-spacing: 0.5px;
    margin-bottom: 0.75rem;
}

/* ── Instruction card ── */
.instruction-card {
    background: #161b22;
    border: 1px solid #21262d;
    border-left: 3px solid #00c8c8;
    border-radius: 8px;
    padding: 1.25rem 1.5rem;
    margin-bottom: 1.5rem;
}
.instruction-card h3 {
    font-family: 'Space Grotesk', sans-serif;
    font-size: 0.85rem;
    font-weight: 700;
    color: #00c8c8;
    letter-spacing: 1px;
    text-transform: uppercase;
    margin: 0 0 0.75rem;
}
.instruction-card ol {
    margin: 0;
    padding-left: 1.2rem;
    color: #8b949e;
    font-size: 0.88rem;
    line-height: 1.7;
}
.instruction-card li span { color: #e6edf3; }

/* ── Metric cards ── */
.metric-grid {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 1rem;
    margin-bottom: 2rem;
}
.metric-card {
    background: #161b22;
    border: 1px solid #21262d;
    border-radius: 10px;
    padding: 1.25rem 1.5rem;
    position: relative;
    overflow: hidden;
}
.metric-card::before {
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 3px;
}
.metric-card.total::before  { background: #58a6ff; }
.metric-card.bully::before  { background: #f85149; }
.metric-card.safe::before   { background: #3fb950; }
.metric-card.rate::before   { background: #d29922; }
.metric-card .mc-label {
    font-size: 0.75rem;
    font-weight: 600;
    letter-spacing: 0.8px;
    text-transform: uppercase;
    color: #8b949e;
    margin-bottom: 0.5rem;
}
.metric-card .mc-value {
    font-family: 'Space Grotesk', sans-serif;
    font-size: 2.2rem;
    font-weight: 700;
    line-height: 1;
}
.metric-card.total  .mc-value { color: #58a6ff; }
.metric-card.bully  .mc-value { color: #f85149; }
.metric-card.safe   .mc-value { color: #3fb950; }
.metric-card.rate   .mc-value { color: #d29922; }
.metric-card .mc-sub {
    font-size: 0.78rem;
    color: #8b949e;
    margin-top: 0.3rem;
}

/* ── Section headings ── */
.section-heading {
    font-family: 'Space Grotesk', sans-serif;
    font-size: 1rem;
    font-weight: 700;
    color: #e6edf3;
    letter-spacing: 0.3px;
    margin: 2rem 0 1rem;
    padding-bottom: 0.5rem;
    border-bottom: 1px solid #21262d;
    display: flex;
    align-items: center;
    gap: 0.5rem;
}

/* ── Chart container ── */
.chart-box {
    background: #161b22;
    border: 1px solid #21262d;
    border-radius: 10px;
    padding: 1.25rem;
}

/* ── Emotion pill ── */
.emotion-pill {
    display: inline-block;
    padding: 0.2rem 0.8rem;
    border-radius: 999px;
    font-size: 0.8rem;
    font-weight: 600;
}

/* ── Info box ── */
.info-box {
    background: rgba(0, 200, 200, 0.06);
    border: 1px solid rgba(0, 200, 200, 0.2);
    border-radius: 8px;
    padding: 0.9rem 1.1rem;
    font-size: 0.85rem;
    color: #8b949e;
    margin: 1rem 0;
}

/* ── Streamlit overrides ── */
div[data-testid="stFileUploader"] {
    background: #161b22;
    border: 1px dashed #30363d;
    border-radius: 10px;
    padding: 1rem;
}
div[data-testid="stFileUploader"] label { color: #8b949e !important; }
div[data-testid="stFileUploader"] button {
    background: #21262d !important;
    border: 1px solid #30363d !important;
    color: #e6edf3 !important;
    border-radius: 6px !important;
}
.stButton > button {
    background: #00c8c8 !important;
    color: #0d1117 !important;
    border: none !important;
    border-radius: 6px !important;
    font-weight: 600 !important;
    font-size: 0.9rem !important;
    padding: 0.6rem 1.4rem !important;
    transition: opacity 0.15s !important;
}
.stButton > button:hover { opacity: 0.88 !important; }
div[data-testid="stDataFrame"] {
    background: #161b22 !important;
    border: 1px solid #21262d !important;
    border-radius: 8px !important;
}
.stTextInput > div > div {
    background: #161b22 !important;
    border-color: #30363d !important;
    color: #e6edf3 !important;
    border-radius: 6px !important;
}
.stAlert { border-radius: 8px !important; }
</style>
""", unsafe_allow_html=True)

# ──────────────────────────────────────────────────────────────────────────────
# STOPWORDS CONFIGURATION  ← insert your values here
# ──────────────────────────────────────────────────────────────────────────────
try:
    from nltk.corpus import stopwords
    nltk_stopwords = set(stopwords.words('english'))
except Exception:
    nltk_stopwords = set()

# ── Customise these two sets to match what was used during training ──────────
all_stopwords: set = nltk_stopwords   # e.g. nltk_stopwords | {"additional", "words"}
keep_words: set    = {"not", "no", "never", "hate", "against"}  # words to keep even if stopwords

# ──────────────────────────────────────────────────────────────────────────────
# PREPROCESSING
# ──────────────────────────────────────────────────────────────────────────────
def preprocess_text(text: str):
    """
    Cleans a raw tweet through the following steps:
    1. Lowercase all characters
    2. Remove URLs
    3. Remove @mentions
    4. Remove hashtag symbols
    5. Remove emojis and non-ASCII characters
    6. Remove punctuation and special characters
    7. Tokenize words
    8. Remove stopwords
    9. Remove short words
    """
    if not isinstance(text, str):
        return "", []

    text = text.lower()
    text = re.sub(r'http\S+|www\S+|https\S+', '', text)
    text = re.sub(r'@\w+', '', text)
    text = re.sub(r'#', '', text)
    text = text.encode('ascii', 'ignore').decode('ascii')
    text = re.sub(r'[^a-z\s]', '', text)
    text = re.sub(r'\s+', ' ', text).strip()

    tokens = word_tokenize(text)
    tokens = [w for w in tokens if w not in all_stopwords or w in keep_words]
    tokens = [w for w in tokens if len(w) > 1]

    cleaned_text = ' '.join(tokens)
    return cleaned_text, tokens


# ──────────────────────────────────────────────────────────────────────────────
# MODEL LOADING
# ──────────────────────────────────────────────────────────────────────────────
@st.cache_resource(show_spinner=False)
def load_models():
    """Load pre-trained models and vectorizers from disk."""
    bully_model      = joblib.load("lr_bully_model.pkl")
    emotion_model    = joblib.load("lr_emotion_model.pkl")
    bully_vectorizer = joblib.load("bully_tfidf_vectorizer.pkl")
    emotion_vectorizer = joblib.load("emotion_tfidf_vectorizer.pkl")
    return bully_model, emotion_model, bully_vectorizer, emotion_vectorizer


# ──────────────────────────────────────────────────────────────────────────────
# PREDICTION
# ──────────────────────────────────────────────────────────────────────────────
def predict_bully(text: str, model, vectorizer) -> str:
    """Return 'Yes' or 'No' for a single preprocessed text."""
    vec = vectorizer.transform([text])
    pred = model.predict(vec)[0]
    # Handle both numeric (1/0) and string labels
    return "Yes" if str(pred) in ("1", "yes", "Yes", "bully", "True") else "No"


def predict_emotion(text: str, model, vectorizer) -> str:
    """Return the predicted emotion label for a single preprocessed text."""
    vec = vectorizer.transform([text])
    return model.predict(vec)[0]


def analyze_dataframe(df: pd.DataFrame, bully_model, emotion_model,
                      bully_vectorizer, emotion_vectorizer) -> pd.DataFrame:
    """Run full analysis pipeline on the DataFrame and return enriched results."""
    results = []
    progress = st.progress(0, text="Analysing comments…")
    total = len(df)

    for i, row in df.iterrows():
        username = str(row["Username"]).strip()
        comment  = str(row["Comment"]).strip()

        cleaned, _ = preprocess_text(comment)

        if cleaned:
            bully = predict_bully(cleaned, bully_model, bully_vectorizer)
        else:
            bully = "No"

        if bully == "Yes" and cleaned:
            emotion = predict_emotion(cleaned, emotion_model, emotion_vectorizer)
        else:
            emotion = "N/A"

        results.append({
            "Username": username,
            "Comment":  comment,
            "Bully":    bully,
            "Emotion":  emotion,
        })
        progress.progress(int((i + 1) / total * 100),
                          text=f"Analysing comments… {i + 1}/{total}")

    progress.empty()
    return pd.DataFrame(results)


# ──────────────────────────────────────────────────────────────────────────────
# VALIDATION
# ──────────────────────────────────────────────────────────────────────────────
def validate_file(df: pd.DataFrame) -> tuple[bool, str]:
    """Check required columns exist and the file has at least one data row."""
    required = {"Username", "Comment"}
    missing = required - set(df.columns)
    if missing:
        return False, f"Missing column(s): **{', '.join(missing)}**. Please check your file."
    if df.empty:
        return False, "The uploaded file contains no data rows."
    return True, ""


# ──────────────────────────────────────────────────────────────────────────────
# EXCEL REPORT GENERATION
# ──────────────────────────────────────────────────────────────────────────────
def build_excel_report(results: pd.DataFrame) -> bytes:
    """Generate a styled Excel workbook with two sheets and return as bytes."""
    wb = Workbook()

    # ── Palette ──────────────────────────────────────────────────────────────
    NAVY   = "0D1117"
    TEAL   = "00C8C8"
    DARK   = "161B22"
    MID    = "21262D"
    WHITE  = "E6EDF3"
    RED    = "F85149"
    GREEN  = "3FB950"
    YELLOW = "D29922"
    GREY   = "8B949E"

    header_font  = Font(name="Calibri", bold=True, color=WHITE, size=11)
    body_font    = Font(name="Calibri", color=WHITE, size=10)
    title_font   = Font(name="Calibri", bold=True, color=TEAL, size=14)
    label_font   = Font(name="Calibri", bold=True, color=GREY, size=10)
    value_font   = Font(name="Calibri", bold=True, color=WHITE, size=12)

    header_fill  = PatternFill("solid", fgColor=MID)
    dark_fill    = PatternFill("solid", fgColor=DARK)
    navy_fill    = PatternFill("solid", fgColor=NAVY)
    teal_fill    = PatternFill("solid", fgColor="0B3A3A")

    thin_side  = Side(style='thin', color=MID)
    thin_border = Border(left=thin_side, right=thin_side,
                         top=thin_side, bottom=thin_side)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # ── Sheet 1 : Detailed Results ────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Detailed Results"
    ws1.sheet_view.showGridLines = False
    ws1.freeze_panes = "A3"

    ws1.row_dimensions[1].height = 36
    ws1.row_dimensions[2].height = 24

    # Title row
    ws1.merge_cells("A1:D1")
    c = ws1["A1"]
    c.value = "Cyberbullying Detection — Detailed Results"
    c.font  = Font(name="Calibri", bold=True, color=TEAL, size=14)
    c.fill  = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Header row
    headers = ["Username", "Comment", "Bully", "Emotion"]
    col_widths = [22, 60, 12, 16]
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws1.cell(row=2, column=col_idx, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = thin_border
        ws1.column_dimensions[get_column_letter(col_idx)].width = w

    # Data rows
    for r_idx, row in enumerate(results.itertuples(index=False), start=3):
        ws1.row_dimensions[r_idx].height = 20
        for c_idx, val in enumerate(row, start=1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=val)
            cell.font   = body_font
            cell.fill   = dark_fill if r_idx % 2 == 0 else PatternFill("solid", fgColor="0F1318")
            cell.border = thin_border
            cell.alignment = center if c_idx in (3, 4) else left

            # Colour-code Bully & Emotion columns
            if c_idx == 3:
                if val == "Yes":
                    cell.font = Font(name="Calibri", bold=True, color=RED, size=10)
                else:
                    cell.font = Font(name="Calibri", bold=True, color=GREEN, size=10)
            if c_idx == 4 and val != "N/A":
                cell.font = Font(name="Calibri", color=TEAL, size=10)

    # ── Sheet 2 : Dashboard Summary ───────────────────────────────────────────
    ws2 = wb.create_sheet("Dashboard Summary")
    ws2.sheet_view.showGridLines = False

    # Metrics
    total_comments   = len(results)
    bully_comments   = (results["Bully"] == "Yes").sum()
    nonbully_comments = total_comments - bully_comments
    bully_rate       = (bully_comments / total_comments * 100) if total_comments else 0

    bully_df = results[results["Bully"] == "Yes"]
    most_common_emotion = (
        bully_df["Emotion"].value_counts().idxmax()
        if not bully_df.empty else "N/A"
    )

    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 22

    def styled_row(ws, row, label, value, val_color):
        r = ws.cell(row=row, column=1, value=label)
        r.font      = label_font
        r.fill      = dark_fill
        r.alignment = left
        r.border    = thin_border

        v = ws.cell(row=row, column=2, value=value)
        v.font      = Font(name="Calibri", bold=True, color=val_color, size=12)
        v.fill      = dark_fill
        v.alignment = center
        v.border    = thin_border
        ws.row_dimensions[row].height = 28

    # Title
    ws2.merge_cells("A1:B1")
    t = ws2["A1"]
    t.value     = "Dashboard Summary"
    t.font      = title_font
    t.fill      = PatternFill("solid", fgColor=NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 36

    ws2.merge_cells("A2:B2")  # spacer

    styled_row(ws2, 3,  "Total Comments Analysed",  total_comments,    "58A6FF")
    styled_row(ws2, 4,  "Bullying Comments",          bully_comments,    RED)
    styled_row(ws2, 5,  "Non-Bullying Comments",      nonbully_comments, GREEN)
    styled_row(ws2, 6,  "Cyberbullying Rate (%)",     f"{bully_rate:.1f}%", YELLOW)
    styled_row(ws2, 7,  "Dominant Emotion",        most_common_emotion, TEAL)

    # ── Serialise ────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ──────────────────────────────────────────────────────────────────────────────
# DASHBOARD RENDERING
# ──────────────────────────────────────────────────────────────────────────────
EMOTION_COLOURS = {
    "Anger":    "#f85149",
    "Sadness":  "#58a6ff",
    "Fear":     "#d29922",
    "Disgust":  "#bc8cff",
    "Surprise": "#3fb950",
    "Neutral":  "#8b949e",
    "Happy":    "#00c8c8",
    "Joy":      "#00c8c8",
}

def render_metrics(results: pd.DataFrame):
    total   = len(results)
    bully   = (results["Bully"] == "Yes").sum()
    nonbully = total - bully
    rate    = (bully / total * 100) if total else 0

    st.markdown(f"""
    <div class="metric-grid">
        <div class="metric-card total">
            <div class="mc-label">Total Comments</div>
            <div class="mc-value">{total}</div>
            <div class="mc-sub">Analysed in this batch</div>
        </div>
        <div class="metric-card bully">
            <div class="mc-label">Bullying Comments</div>
            <div class="mc-value">{bully}</div>
            <div class="mc-sub">Flagged for review</div>
        </div>
        <div class="metric-card safe">
            <div class="mc-label">Non-Bullying</div>
            <div class="mc-value">{nonbully}</div>
            <div class="mc-sub">No indicators found</div>
        </div>
        <div class="metric-card rate">
            <div class="mc-label">Bullying Rate</div>
            <div class="mc-value">{rate:.1f}%</div>
            <div class="mc-sub">Of all comments</div>
        </div>
    </div>
    """, unsafe_allow_html=True)


def render_emotion_section(results: pd.DataFrame):
    bully_df = results[results["Bully"] == "Yes"]
    if bully_df.empty:
        st.info("No bullying comments were detected — emotion analysis is not applicable.")
        return

    emotion_counts = bully_df["Emotion"].value_counts().reset_index()
    emotion_counts.columns = ["Emotion", "Count"]

    colours = [EMOTION_COLOURS.get(e, "#8b949e") for e in emotion_counts["Emotion"]]

    col1, col2 = st.columns([2, 1])

    with col1:
        st.markdown('<div class="chart-box">', unsafe_allow_html=True)
        fig = px.pie(
            emotion_counts,
            names="Emotion",
            values="Count",
            color="Emotion",
            color_discrete_map=EMOTION_COLOURS,
            hole=0.45,
        )
        fig.update_traces(
            textinfo="label+percent",
            textfont_color="#e6edf3",
            marker=dict(line=dict(color="#0d1117", width=2)),
        )
        fig.update_layout(
            paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(0,0,0,0)",
            font_color="#e6edf3",
            legend=dict(
                bgcolor="rgba(0,0,0,0)",
                font=dict(color="#8b949e", size=11),
            ),
            margin=dict(t=20, b=20, l=20, r=20),
            showlegend=True,
        )
        st.plotly_chart(fig, use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)

    with col2:
        most_common = emotion_counts.iloc[0]["Emotion"]
        pct = emotion_counts.iloc[0]["Count"] / len(bully_df) * 100

        colour = EMOTION_COLOURS.get(most_common, "#8b949e")
        st.markdown(f"""
        <div class="chart-box" style="height:100%; display:flex; flex-direction:column;
             justify-content:center; align-items:center; text-align:center; gap:0.75rem;">
            <div style="font-size:0.75rem;font-weight:600;letter-spacing:1px;
                 text-transform:uppercase;color:#8b949e;">Most Common Emotion</div>
            <div style="font-family:'Space Grotesk',sans-serif;font-size:2.4rem;
                 font-weight:700;color:{colour};">{most_common}</div>
            <div style="font-size:0.85rem;color:#8b949e;">
                {pct:.1f}% of bullying comments
            </div>
            <hr style="border:0;border-top:1px solid #21262d;width:80%;margin:0.5rem 0;">
            <div style="font-size:0.75rem;color:#8b949e;">
                Total bullying comments<br>
                <span style="color:#f85149;font-size:1.3rem;font-weight:700;">{len(bully_df)}</span>
            </div>
        </div>
        """, unsafe_allow_html=True)


def render_risk_users(results: pd.DataFrame):
    grouped = (
        results.groupby("Username")
        .agg(
            total_comments=("Comment", "count"),
            bully_comments=("Bully", lambda x: (x == "Yes").sum()),
        )
        .reset_index()
    )
    # Keep only users with ≥ 3 comments
    grouped = grouped[grouped["total_comments"] >= 3].copy()
    grouped["Bully Rate (%)"] = (
        grouped["bully_comments"] / grouped["total_comments"] * 100
    ).round(1)
    grouped = grouped.rename(columns={
        "Username": "Username",
        "total_comments": "Total Comments",
        "bully_comments": "Bullying Comments",
    })
    grouped = grouped.sort_values("Bully Rate (%)", ascending=False).reset_index(drop=True)

    if grouped.empty:
        st.info("No users with 3 or more comments found. Increase the dataset size for user-level analysis.")
        return

    # Highlight rows with high bully rate
    def highlight_rate(val):
        if val >= 60:
            return "color: #f85149; font-weight: 600;"
        elif val >= 30:
            return "color: #d29922; font-weight: 600;"
        return "color: #3fb950; font-weight: 600;"

    styled = grouped.style.applymap(highlight_rate, subset=["Bully Rate (%)"])
    st.dataframe(
        styled,
        use_container_width=True,
        hide_index=True,
        column_config={
            "Bully Rate (%)": st.column_config.NumberColumn(format="%.1f%%"),
        },
    )

    st.markdown(
        '<div class="info-box">ℹ️ Users are ranked by bullying rate. '
        'This table is a decision-support tool — it does <strong>not</strong> label any user as a cyberbully.</div>',
        unsafe_allow_html=True,
    )


def render_detailed_results(results: pd.DataFrame):
    search = st.text_input("🔍  Filter by keyword (username, comment, emotion…)", "")

    df_show = results.copy()
    if search:
        mask = df_show.apply(
            lambda row: row.astype(str).str.contains(search, case=False).any(), axis=1
        )
        df_show = df_show[mask]

    def colour_bully(val):
        return "color: #f85149; font-weight:600;" if val == "Yes" else "color: #3fb950;"

    def colour_emotion(val):
        c = EMOTION_COLOURS.get(val, "#8b949e")
        return f"color: {c};" if val != "N/A" else "color: #8b949e;"

    styled = (
        df_show.style
        .applymap(colour_bully,   subset=["Bully"])
        .applymap(colour_emotion, subset=["Emotion"])
    )

    st.dataframe(styled, use_container_width=True, hide_index=True)
    st.caption(f"Showing {len(df_show):,} of {len(results):,} rows")


# ──────────────────────────────────────────────────────────────────────────────
# MAIN APP
# ──────────────────────────────────────────────────────────────────────────────
def main():
    # ── Header ───────────────────────────────────────────────────────────────
    st.markdown("""
    <div class="cg-header">
        <div class="cg-badge">🛡️ CYBERGUARD</div>
        <h1>Cyberbullying Detection &amp; Emotion Analysis</h1>
        <p>Upload social media comments to identify potential bullying patterns and emotional indicators.</p>
    </div>
    """, unsafe_allow_html=True)

    # ── Instructions ─────────────────────────────────────────────────────────
    st.markdown("""
    <div class="instruction-card">
        <h3>📋 How to Prepare Your File</h3>
        <ol>
            <li>Collect comments from a social media platform.</li>
            <li>Save the comments into <span>CSV</span> or <span>Excel (.xlsx)</span> format.</li>
            <li>Include the following columns: <span>Username</span> and <span>Comment</span>.</li>
            <li>Upload the file using the uploader below.</li>
        </ol>
    </div>
    """, unsafe_allow_html=True)

    # ── Model loading ─────────────────────────────────────────────────────────
    with st.spinner("Loading models…"):
        try:
            bully_model, emotion_model, bully_vectorizer, emotion_vectorizer = load_models()
        except FileNotFoundError as e:
            st.error(
                f"⚠️ Model file not found: `{e.filename}`. "
                "Ensure all four `.pkl` files are in the same directory as `app.py`."
            )
            return

    # ── File upload ───────────────────────────────────────────────────────────
    uploaded = st.file_uploader(
        "Upload your comments file (CSV or Excel)",
        type=["csv", "xlsx"],
        label_visibility="collapsed",
    )

    if uploaded is None:
        st.markdown("""
        <div class="info-box">
            📂 Drag and drop a <strong>CSV</strong> or <strong>Excel (.xlsx)</strong> file above to begin analysis.
        </div>
        """, unsafe_allow_html=True)
        return

    # ── Parse file ────────────────────────────────────────────────────────────
    try:
        if uploaded.name.endswith(".csv"):
            df = pd.read_csv(uploaded)
        else:
            df = pd.read_excel(uploaded, engine="openpyxl")
    except Exception as e:
        st.error(f"Could not read the uploaded file: {e}")
        return

    # ── Validate ──────────────────────────────────────────────────────────────
    valid, msg = validate_file(df)
    if not valid:
        st.error(f"❌ {msg}")
        return

    df = df[["Username", "Comment"]].dropna(subset=["Comment"]).reset_index(drop=True)

    st.success(f"✅ File loaded — **{len(df):,}** comments ready for analysis.")

    if st.button("▶  Run Analysis"):
        # ── Run pipeline ──────────────────────────────────────────────────────
        results = analyze_dataframe(
            df, bully_model, emotion_model, bully_vectorizer, emotion_vectorizer
        )
        st.session_state["results"] = results

    # ── Render results if available ───────────────────────────────────────────
    if "results" not in st.session_state:
        return

    results: pd.DataFrame = st.session_state["results"]

    # Metrics
    st.markdown('<div class="section-heading">📊 Overview</div>', unsafe_allow_html=True)
    render_metrics(results)

    # Emotion analysis
    st.markdown('<div class="section-heading">😶 Emotion Analysis (Bullying Comments Only)</div>',
                unsafe_allow_html=True)
    render_emotion_section(results)

    # Risk users
    st.markdown('<div class="section-heading">⚠️ Potential Risk Users</div>',
                unsafe_allow_html=True)
    render_risk_users(results)

    # Detailed results
    st.markdown('<div class="section-heading">📄 Detailed Results</div>',
                unsafe_allow_html=True)
    render_detailed_results(results)

    # Export
    st.markdown("---")
    excel_bytes = build_excel_report(results)
    st.download_button(
        label="⬇️  Download Analysis Report",
        data=excel_bytes,
        file_name="analysis_report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    main()
